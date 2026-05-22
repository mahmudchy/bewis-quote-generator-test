import streamlit as st
import pandas as pd
import os
import datetime
import requests
from bs4 import BeautifulSoup
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.cell.cell import MergedCell, Cell
from openpyxl.styles import Alignment, Border, Side

# --- 1. DATA LOADING ---
@st.cache_data
def load_all_models():
    all_data = []
    files = [f for f in os.listdir('.') if f.endswith(('.csv', '.xlsx'))]
    for file in files:
        try:
            if any(x in file.lower() for x in ["template", "requirements"]): continue
            df_raw = pd.read_csv(file, header=None).fillna('') if file.endswith('.csv') else pd.read_excel(file, header=None).fillna('')
            model_col_idx = -1
            header_row = 0
            for r_idx in range(min(len(df_raw), 25)):
                row_vals = [str(v).strip().lower() for v in df_raw.iloc[r_idx]]
                if 'model' in row_vals or 'bewis no' in row_vals:
                    model_col_idx = row_vals.index('model') if 'model' in row_vals else row_vals.index('bewis no')
                    header_row = r_idx
                    break
            if model_col_idx != -1:
                df = pd.read_csv(file, header=header_row).fillna('') if file.endswith('.csv') else pd.read_excel(file, header=header_row).fillna('')
                df.columns = [str(c).strip() for c in df.columns]
                m_col = df.columns[model_col_idx]
                for _, row in df.iterrows():
                    m_name = str(row[m_col]).strip()
                    if not m_name or m_name.lower() in ['model', 'nan']: continue
                    specs = []
                    for col in df.columns:
                        if any(k in col.lower() for k in ['accuracy', 'range', 'axis', 'output']):
                            val = str(row[col]).strip()
                            if val and val.lower() != 'nan': specs.append(f"{col}: {val}")
                    all_data.append({"Model": m_name, "Specs": "\n".join(specs)})
        except: continue
    return pd.DataFrame(all_data)

# --- 2. IMAGE SCRAPER ---
def get_bw_sensing_image(model_name):
    base = "https://www.bw-sensing.com"
    search_url = f"{base}/search.html?q={model_name}"
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        res = requests.get(search_url, timeout=7, headers=headers)
        soup = BeautifulSoup(res.text, 'html.parser')
        link = soup.select_one('.product-list a') or soup.select_one('.pro_list a')
        if not link: return None
        detail_url = base + link['href'] if link['href'].startswith('/') else link['href']
        d_res = requests.get(detail_url, timeout=7, headers=headers)
        dsoup = BeautifulSoup(d_res.text, 'html.parser')
        img_tag = dsoup.select_one('.product-info img') or dsoup.select_one('.left-img img')
        if img_tag:
            src = img_tag.get('data-original') or img_tag.get('src')
            if src: return src if src.startswith('http') else base + src
    except: return None
    return None

# --- 3. THE "MERGE-PROOF" WRITER ---
def ultra_safe_write(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for m_range in ws.merged_cells.ranges:
            if cell.coordinate in m_range:
                ws.cell(row=m_range.min_row, column=m_range.min_col).value = value
                return
    cell.value = value

# --- 4. UI SETUP ---
st.set_page_config(layout="wide", page_title="BWS Quote Gen")

st.markdown("""<style>input[type=number]::-webkit-inner-spin-button, input[type=number]::-webkit-outer-spin-button { -webkit-appearance: none; margin: 0; }</style>""", unsafe_allow_html=True)

model_db = load_all_models()
if 'rows' not in st.session_state: st.session_state.rows = [{"model": ""}]

# ... (Sidebar and Input Logic remain the same) ...

# --- 5. EXPORT ---
if st.button("🚀 Export to Excel"):
    if os.path.exists('template.xlsx') and final_data:
        wb = load_workbook('template.xlsx')
        ws = wb.active
        
        # Capture reference border from the first row of the template (e.g., cell A17)
        ref_cell = ws.cell(row=17, column=1)
        ref_border = ref_cell.border
        
        start_row = 17
        for idx, block in enumerate(final_data):
            cur_top = start_row + (idx * 3)
            
            # Write data
            ultra_safe_write(ws, cur_top, 4, block['model'])
            ultra_safe_write(ws, cur_top, 9, block['specs'])
            ultra_safe_write(ws, cur_top, 1, "ALL")
            
            # Apply reference borders to the 3-row block
            for r in range(cur_top, cur_top + 3):
                for c in range(1, 10): # Assuming 9 columns
                    cell = ws.cell(row=r, column=c)
                    cell.border = ref_border
                    cell.alignment = Alignment(vertical='center', horizontal='center', wrapText=True)

            # Tiers & Image logic...
            # (Keep your existing Tiers loop and Image logic here)
            
        out = BytesIO()
        wb.save(out)
        st.download_button("📥 Download Final Excel", out.getvalue(), f"{quote_id}.xlsx")
